"""XLSX: the format data actually arrives in.

`PROPOSAL.md` §6 calls it "extremely common as the real-world interchange
format" and puts it in Phase 1. It is also the second implementation of the
reader interface, which is what turns that interface from a design into a
proven one — everything a spreadsheet and a text file share now lives in
`grid.py`, and what is left here is what only a workbook has.

**What a workbook removes.** There is no delimiter to guess and no encoding to
sniff; cells arrive already separated. `Detection.delimiter` is therefore fixed
and offers no alternatives, which is the honest way to say "nothing to decide"
in a shape built for a format where there was.

**What a workbook adds.** A sheet to choose, and rows above the header that no
text file has: a title, a blank line, an export banner. Both are handled here.

**What a workbook does not remove.** Numbers stored as text — `'1,5` typed into
a cell, or a column imported from a CSV and never converted — are so common
that trusting the cell type is how a reader imports a column of blanks and
reports success. Every cell is rendered to a string and put through the same
detection a text file gets, so a decimal comma is caught in a spreadsheet
exactly as it is in a CSV.
"""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from chemometrics_workbench.readers import (
    Choice,
    Detection,
    Imported,
    ReaderError,
    grid,
    source_file,
)

NAME = "xlsx"
VERSION = "1"
SUFFIXES = (".xlsx", ".xlsm")

#: A workbook has no delimiter. The field stays in the shape because the import
#: preview is one contract across every reader; it just has nothing to offer.
NOT_APPLICABLE = "n/a"

#: A row with fewer than this many filled cells above the data is a title, a
#: banner or a stray note rather than a header.
MIN_HEADER_CELLS = 2


def sniff(path: Path) -> Detection:
    """Open the workbook, find the table in it, and report what it looks like."""
    sheets = _sheet_names(path)
    return _detect(path, sheets[0], sheets)


def read(path: Path, detection: Detection) -> Imported:
    """Read the chosen sheet according to `detection`."""
    sheet = detection.sheet.value if detection.sheet else _sheet_names(path)[0]
    rows = _grid(path, sheet)
    decimal = detection.decimal.value
    header, body = grid.split_header(rows, decimal)
    if not body:
        raise ReaderError(f"{path.name}, sheet {sheet!r}, has a header but no data rows.")

    provenance = source_file(path, NAME, VERSION)
    if detection.orientation.value == "samples_in_columns":
        return grid.read_transposed(provenance, header, body, decimal, detection)
    return grid.read_rows(provenance, header, body, decimal, detection)


def head(path: Path, detection: Detection) -> dict[str, Any]:
    """The first rows of the chosen sheet, for the preview's table."""
    sheet = detection.sheet.value if detection.sheet else _sheet_names(path)[0]
    rows = _grid(path, sheet)
    header, body = grid.split_header(rows, detection.decimal.value)
    return grid.head_payload(
        path.name, header, body, detection.decimal.value, detection.orientation.value
    )


def resniff(path: Path, detection: Detection) -> Detection:
    """Detect again, because a corrected sheet is a different table.

    Correcting a delimiter changes how the same bytes are cut up; correcting a
    sheet changes which bytes are being read at all, so the sample count, the
    axis and every column classification have to be worked out afresh. The
    generic `preview` and `read` call this when a reader offers it.
    """
    sheets = _sheet_names(path)
    sheet = detection.sheet.value if detection.sheet else sheets[0]
    return _detect(path, sheet, sheets)


def _detect(path: Path, sheet: str, sheets: list[str]) -> Detection:
    rows = _grid(path, sheet)
    decimal = _detect_decimal(rows)
    header, body = grid.split_header(rows, decimal)
    if not body:
        raise ReaderError(f"{path.name}, sheet {sheet!r}, has a header but no data rows.")

    orientation = grid.detect_orientation(header, body, decimal)
    sheet_choice = Choice(sheet, tuple(name for name in sheets if name != sheet))
    common = {
        "delimiter": Choice(NOT_APPLICABLE),
        "decimal": Choice(decimal, tuple(d for d in (".", ",") if d != decimal)),
        "sheet": sheet_choice,
        "correctable": ("decimal", "orientation", "sheet"),
    }

    if orientation == "samples_in_columns":
        axis_values = [
            grid.parse_number(row[0], decimal, path.name, line) for line, row in enumerate(body)
        ]
        axis, reconstructed, note = grid.axis_from_values(axis_values, len(axis_values))
        return Detection(
            orientation=Choice("samples_in_columns", ("samples_in_rows",)),
            n_samples=len(header or []) - 1,
            n_variables=len(body),
            axis=axis,
            axis_reconstructed=reconstructed,
            axis_note=note,
            **common,  # type: ignore[arg-type]
        )

    columns = grid.classify(header, body, decimal)
    if not columns["spectra"]:
        raise ReaderError(grid.no_spectra_message(f"{path.name}, sheet {sheet!r}", header, decimal))
    axis, reconstructed, note = grid.axis_from(header, columns["spectra"])
    return Detection(
        orientation=Choice("samples_in_rows", ("samples_in_columns",)),
        n_samples=len(body),
        n_variables=len(columns["spectra"]),
        axis=axis,
        axis_reconstructed=reconstructed,
        axis_note=note,
        metadata_columns=tuple(columns["metadata_names"]),
        targets=tuple(columns["target_names"]),
        discarded=tuple(columns["discarded"]),
        **common,  # type: ignore[arg-type]
    )


def _sheet_names(path: Path) -> list[str]:
    workbook = _open(path)
    try:
        names = list(workbook.sheetnames)
    finally:
        workbook.close()
    if not names:
        raise ReaderError(f"{path.name} has no sheets in it.")
    return names


def _open(path: Path) -> Any:
    """Open read-only, with formulas resolved to their last cached value.

    `data_only=True` returns what Excel computed and saved. A workbook written
    by something that does not cache formula results gives `None` there, and
    those cells are then reported as empty rather than as zero — a zero nobody
    measured is worse than a gap that is visible.
    """
    import zipfile

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        return load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as error:
        raise ReaderError(
            f"{path.name} is not a workbook this build can open. An .xlsx file is a zip "
            f"archive, and this one is not one: {error}"
        ) from error
    except (OSError, KeyError, ValueError) as error:
        raise ReaderError(f"cannot open {path.name}: {error}") from error


def _grid(path: Path, sheet: str) -> list[list[str]]:
    """The sheet as a rectangle of strings, with the packaging stripped off.

    Trailing empty rows and columns go, because a workbook's declared dimensions
    are routinely far larger than its data and a reader that believes them
    imports a thousand blank samples. Leading rows with fewer than
    `MIN_HEADER_CELLS` filled cells go too: that is a title, a banner or a note
    above the table, and it is what shifts a naive reader's columns by one.
    """
    workbook = _open(path)
    try:
        if sheet not in workbook.sheetnames:
            raise ReaderError(
                f"{path.name} has no sheet named {sheet!r}. It has: "
                f"{', '.join(workbook.sheetnames)}."
            )
        rows = [
            [_text(cell) for cell in row] for row in workbook[sheet].iter_rows(values_only=True)
        ]
    finally:
        workbook.close()

    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        raise ReaderError(f"{path.name}, sheet {sheet!r}, is empty.")

    while rows and sum(1 for cell in rows[0] if cell) < MIN_HEADER_CELLS:
        rows.pop(0)
    if not rows:
        raise ReaderError(
            f"{path.name}, sheet {sheet!r}, has no row with {MIN_HEADER_CELLS} or more "
            "filled cells, so there is no table in it."
        )

    width = max(len(row) for row in rows)
    rows = [row + [""] * (width - len(row)) for row in rows]
    keep = [i for i in range(width) if any(row[i] for row in rows)]
    return [[row[i] for i in keep] for row in rows]


def _text(value: object) -> str:
    """Render a cell as the string a text reader would have seen.

    Dates become ISO text rather than the serial number underneath them, so a
    date column reads as metadata instead of as a plausible absorbance. Floats
    keep `repr`'s round-tripping, so nothing is lost on the way through a string
    and back.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, float):
        return repr(value)
    return str(value).strip()


def _detect_decimal(rows: list[list[str]]) -> str:
    """A workbook can still hold its numbers as `1,5` text. Ask the same question."""
    european = sum(1 for row in rows for cell in row if grid.is_number(cell, ",") and "," in cell)
    plain = sum(1 for row in rows for cell in row if grid.is_number(cell, "."))
    return "," if european > plain else "."
